import openpyxl
import json
from io import BytesIO
from typing import Dict, Any

def parse_assessment_excel(file_content: bytes) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(BytesIO(file_content))
    
    score_ws = wb.get_sheet_by_name('分数统计表') or wb.active
    score_data = parse_score_table(score_ws)
    
    behavior_ws = wb.get_sheet_by_name('行为记录表')
    behavior_data = parse_behavior_table(behavior_ws) if behavior_ws else []
    
    return {
        "score_table": score_data,
        "behavior_table": behavior_data
    }

def parse_score_table(ws) -> list:
    data = []
    headers = []
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(cell) if cell else "" for cell in row]
            continue
        
        if not any(row):
            continue
        
        row_data = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_data[headers[j]] = cell
        data.append(row_data)
    
    return data

def parse_behavior_table(ws) -> list:
    data = []
    headers = []
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(cell) if cell else "" for cell in row]
            continue
        
        if not any(row):
            continue
        
        row_data = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                row_data[headers[j]] = cell
        data.append(row_data)
    
    return data

def create_template_excel() -> bytes:
    wb = openpyxl.Workbook()
    
    score_ws = wb.active
    score_ws.title = "分数统计表"
    score_headers = ["考生ID", "考生姓名", "领导力", "沟通协作", "创新能力", "分析思维"]
    score_ws.append(score_headers)
    score_ws.append(["A1", "张三", "", "", "", ""])
    score_ws.append(["A2", "李四", "", "", "", ""])
    
    behavior_ws = wb.create_sheet("行为记录表")
    behavior_headers = ["考生ID", "考生姓名", "测评环节", "能力维度", "行为描述"]
    behavior_ws.append(behavior_headers)
    behavior_ws.append(["A1", "张三", "BEI", "沟通协作", ""])
    behavior_ws.append(["A1", "张三", "LGD", "领导力", ""])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
