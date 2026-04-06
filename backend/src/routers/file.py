from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from src.database import get_db
from src.models.user import User
from src.routers.auth import get_current_user
from src.models.file import File as FileModel
from src.utils.document import markdown_to_docx, generate_radar_chart
from src.utils.parser import parse_assessment_excel, create_template_excel
import os
import uuid
import aiofiles
from src.config import settings

def parse_file_content(file_path: str, file_type: str) -> str:
    """解析不同类型的文件并返回文本内容"""
    ext = os.path.splitext(file_path)[1].lower()
    
    # Word 文件 (.docx)
    if ext == '.docx' or (file_type and 'word' in file_type.lower()):
        try:
            from docx import Document
            doc = Document(file_path)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return '\n'.join(paragraphs)
        except Exception as e:
            return f"Word文件解析失败: {str(e)}"
    
    # Excel 文件 (.xlsx, .xls)
    elif ext in ['.xlsx', '.xls'] or (file_type and 'excel' in file_type.lower() or 'sheet' in file_type.lower()):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            content_parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                content_parts.append(f"\n--- Sheet: {sheet} ---\n")
                for row in ws.iter_rows(values_only=True):
                    row_text = ' | '.join([str(cell) if cell is not None else '' for cell in row])
                    if row_text.strip():
                        content_parts.append(row_text)
            return '\n'.join(content_parts)
        except Exception as e:
            return f"Excel文件解析失败: {str(e)}"
    
    # PDF 文件
    elif ext == '.pdf' or (file_type and 'pdf' in file_type.lower()):
        try:
            from pypdf import PdfReader
            reader = PdfReader(file_path)
            content_parts = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    content_parts.append(text)
            return '\n'.join(content_parts)
        except Exception as e:
            return f"PDF文件解析失败: {str(e)}"
    
    # 纯文本文件
    else:
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        except Exception as e:
            return f"文件读取失败: {str(e)}"


router = APIRouter(prefix="/api/files", tags=["文件处理"])

@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    file_id = str(uuid.uuid4())
    ext = os.path.splitext(file.filename)[1]
    file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}{ext}")
    
    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    
    async with aiofiles.open(file_path, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    db_file = FileModel(
        user_id=current_user.id,
        name=file.filename,
        type=file.content_type,
        size=len(content),
        file_path=file_path
    )
    db.add(db_file)
    await db.commit()
    await db.refresh(db_file)
    
    return {
        "success": True,
        "data": {
            "id": db_file.id,
            "name": db_file.name,
            "type": db_file.type,
            "size": db_file.size,
            "file_path": db_file.file_path
        }
    }

@router.get("/{file_id}/content")
async def get_file_content(
    file_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(FileModel).where(
            FileModel.id == file_id,
            FileModel.user_id == current_user.id
        )
    )
    file = result.scalar_one_or_none()
    if not file:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    # 使用新的解析函数提取文件内容
    try:
        content = parse_file_content(file.file_path, file.type or "")
        return {"success": True, "data": {"content": content, "type": get_file_type(file.type or "")}}
    except Exception as e:
        return {"success": True, "data": {"content": f"文件解析失败: {str(e)}", "type": "error"}}


def get_file_type(content_type: str) -> str:
    """根据content-type返回文件类型"""
    if 'pdf' in content_type.lower():
        return 'pdf'
    elif 'word' in content_type.lower() or 'document' in content_type.lower():
        return 'word'
    elif 'excel' in content_type.lower() or 'sheet' in content_type.lower():
        return 'excel'
    elif 'text' in content_type.lower():
        return 'text'
    return 'unknown'
@router.post("/parse-assessment")
async def parse_assessment(
    file_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(FileModel).where(
            FileModel.id == file_id,
            FileModel.user_id == current_user.id
        )
    )
    file = result.scalar_one_or_none()
    if not file:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    async with aiofiles.open(file.file_path, 'rb') as f:
        content = await f.read()
    
    parsed = parse_assessment_excel(content)
    
    return {"success": True, "data": parsed}

@router.get("/download-template")
async def download_template():
    content = create_template_excel()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=assessment_template.xlsx"}
    )

@router.get("/download/{file_id}")
async def download_file(
    file_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(FileModel).where(
            FileModel.id == file_id,
            FileModel.user_id == current_user.id
        )
    )
    file = result.scalar_one_or_none()
    if not file:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    async with aiofiles.open(file.file_path, 'rb') as f:
        content = await f.read()
    
    return Response(
        content=content,
        media_type=file.type,
        headers={"Content-Disposition": f"attachment; filename={file.name}"}
    )

@router.post("/convert/markdown-to-docx")
async def convert_markdown_to_docx(
    content: str,
    title: str = "",
    current_user: User = Depends(get_current_user)
):
    docx_content = markdown_to_docx(content, title)
    
    os.makedirs(settings.GENERATED_DIR, exist_ok=True)
    file_path = os.path.join(settings.GENERATED_DIR, f"{uuid.uuid4()}.docx")
    
    with open(file_path, 'wb') as f:
        f.write(docx_content)
    
    return Response(
        content=docx_content,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename={title or 'document'}.docx"}
    )

@router.post("/generate-radar-chart")
async def create_radar_chart(
    scores: dict,
    title: str = "能力雷达图",
    current_user: User = Depends(get_current_user)
):
    chart_base64 = generate_radar_chart(scores, title)
    return {"success": True, "data": {"chart": chart_base64}}
